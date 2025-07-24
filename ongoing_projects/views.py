from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from .models import OngoingProject, ProjectComponentStatus
from .forms import OngoingProjectForm, ProjectComponentStatusFormSet
from django.forms import modelformset_factory
from django.db.models import Q
from django.contrib import messages
import openpyxl
from django.http import HttpResponse
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from django.utils.dateparse import parse_date
from django.db.models import Count
from delivery_notes.models import DeliveryNote
from incoming_material.models import IncomingMaterialReport
from ongoing_projects.models import OngoingProject

@login_required
def list_ongoing_projects(request):
    query = request.GET.get('q', '')
    filter_date = request.GET.get('date')
    projects = OngoingProject.objects.all()
    if query:
        projects = projects.filter(
            Q(code__icontains=query) |
            Q(po_number__icontains=query) |
            Q(client__icontains=query) |
            Q(location__icontains=query) |
            Q(po_description__icontains=query)
        )
    if filter_date:
        date_obj = parse_date(filter_date)  
        if date_obj:
            projects = projects.filter(date=date_obj)
    paginator = Paginator(projects.order_by('-date'), 8)
    page_number = request.GET.get('page')
    projects = paginator.get_page(page_number)

    return render(request, 'ongoing_projects/list.html', {
        'projects': projects
    })

@login_required 
def add_ongoing_project(request):
    component_labels = ['3m Fence', '5m Fence', 'Gate']
    if request.method == 'POST':
        form = OngoingProjectForm(request.POST)
        if form.is_valid():
            project = form.save(commit=False)
            project.status = 'In Progress'  # 🔧 Set default status manually
            project.save()
            count = 1
            while True:
                component_name = request.POST.get(f'component_name_{count}')
                if not component_name:
                    break
                material = request.POST.get(f'material_procurement_{count}', 'Pending')
                fabrication = request.POST.get(f'fabrication_{count}', 'N/A')
                hdg = request.POST.get(f'hdg_{count}', 'N/A')
                coating = request.POST.get(f'powder_coating_{count}', 'N/A')
                erection = request.POST.get(f'erection_{count}', 'N/A')
                ProjectComponentStatus.objects.create(
                    project=project,
                    component=component_name,
                    material_procurement=material,
                    fabrication=fabrication,
                    hdg=hdg,
                    powder_coating=coating,
                    erection=erection
                )
                count += 1
            messages.success(request, "Ongoing Project Successfully Added")  # Add this line
            return redirect('view_ongoing_project', pk=project.pk)
        else:
            print("❌ Form is not valid")
            print(form.errors)
    else:
        form = OngoingProjectForm()

    return render(request, 'ongoing_projects/add.html', {
        'form': form,
        'component_labels': component_labels
    })

@login_required
def view_ongoing_project(request, pk):
    project = get_object_or_404(OngoingProject, pk=pk)
    components = project.components.all()
    return render(request, 'ongoing_projects/view.html', {
        'project': project,
        'components': components
    })

@login_required
def edit_ongoing_project(request, pk):
    project = get_object_or_404(OngoingProject, pk=pk)
    ProjectComponentStatusFormSet = modelformset_factory(
        ProjectComponentStatus,
        fields=('component', 'material_procurement', 'fabrication', 'hdg', 'powder_coating', 'erection'),
        extra=0,
        can_delete=True
    )

    if request.method == 'POST':
        form = OngoingProjectForm(request.POST, instance=project)
        formset = ProjectComponentStatusFormSet(request.POST, queryset=project.components.all())

        if form.is_valid() and formset.is_valid():
            form.save()
            components = formset.save(commit=False)

            for component in components:
                component.project = project
                component.save()

            for obj in formset.deleted_objects:
                obj.delete()
            
            # Add success message
            messages.success(request, f"Project {project.code} updated successfully")
            return redirect('view_ongoing_project', pk=project.pk)

    else:
        form = OngoingProjectForm(instance=project)
        formset = ProjectComponentStatusFormSet(queryset=project.components.all())

    return render(request, 'ongoing_projects/edit.html', {
        'form': form,
        'formset': formset,
        'project': project,
    })

@login_required
def delete_ongoing_project(request, pk):
    project = get_object_or_404(OngoingProject, pk=pk)
    project_code = project.code  # Get the code before deletion
    project.delete()
    messages.success(request, f"Project {project_code} deleted successfully")
    return redirect('ongoing_projects_list')

@login_required
def export_ongoing_projects_to_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    projects = OngoingProject.objects.prefetch_related('components').all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Ongoing Projects"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # Main Table Header
    headers = ["S.No", "CODE", "Client", "Project Location", "P.O Description",
               "P.O Amount", "Budgetary Cost", "Expected Profit", "Down Payment", "Shop Drawing"]
    
    row = 1
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row += 1
    for i, project in enumerate(projects, 1):
        project_data = [
            i, project.code, project.client, project.location, project.po_description,
            project.po_amount, project.budgetary_cost, project.expected_profit,
            project.down_payment, project.shop_drawing
        ]
        for col, val in enumerate(project_data, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = center
            cell.border = border
        row += 1

        # Component Sub-table
        component_headers = ["Component", "Material", "Fabrication", "HDG", "Powder Coating", "Erection"]
        for col, title in enumerate(component_headers, start=1):
            cell = ws.cell(row=row, column=col, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        row += 1

        for c in project.components.all():
            component_data = [
                c.component, c.material_procurement, c.fabrication,
                c.hdg, c.powder_coating, c.erection
            ]
            for col, val in enumerate(component_data, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = center
                cell.border = border
            row += 1

        row += 1  # Extra spacing row

    # Auto column width
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 4

    # Return the response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ongoing_projects.xlsx"'
    wb.save(response)
    return response


@login_required
def export_single_project_excel(request, pk):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from django.http import HttpResponse

    project = get_object_or_404(OngoingProject, pk=pk)
    components = project.components.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Project {project.code}"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Project Info Header Row
    headers = ["Code", "PO No", "Description", "PO Amount", "Client", "Budgetary Cost", "Location", "Expected Profit", "Shop Drawing", "Down Payment"]
    ws.append(headers)

    # Apply style to header
    for col_num, col in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Project Data Row
    ws.append([
        project.code,
        project.po_number,
        project.po_description,
        project.po_amount,
        project.client,
        project.budgetary_cost,
        project.location,
        project.expected_profit,
        project.shop_drawing,
        project.down_payment
    ])
    for col_num in range(1, len(headers)+1):
        ws.cell(row=2, column=col_num).alignment = center

    # Spacing
    ws.append([])

    # Components Table Header
    component_headers = ["Component", "Material", "Fabrication", "HDG", "Powder Coating", "Erection"]
    ws.append(component_headers)

    for col_num, col in enumerate(component_headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Components Data
    row = 5
    for c in components:
        ws.append([
            c.component,
            c.material_procurement,
            c.fabrication,
            c.hdg,
            c.powder_coating,
            c.erection,
        ])
        for col in range(1, 7):
            ws.cell(row=row, column=col).alignment = center
        row += 1

    # Auto column width
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 4

    # Output
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"{project.code}_ongoing_project.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def download_single_project_pdf(request, pk):
    project = get_object_or_404(OngoingProject, pk=pk)
    components = project.components.all()

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    y = height - 50
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, f"Ongoing Project: {project.code}")
    y -= 30

    p.setFont("Helvetica", 12)
    details = [
        ("PO No", project.po_number),
        ("Description", project.po_description),
        ("PO Amount", project.po_amount),
        ("Client", project.client),
        ("Budgetary Cost", project.budgetary_cost),
        ("Location", project.location),
        ("Expected Profit", project.expected_profit),
        ("Shop Drawing", project.shop_drawing),
        ("Down Payment", project.down_payment),
    ]

    for field, val in details:
        p.drawString(50, y, f"{field}: {val}")
        y -= 20

    y -= 20
    p.setFont("Helvetica-Bold", 13)
    p.drawString(50, y, "Component Status:")
    y -= 20

    p.setFont("Helvetica", 11)
    headers = ["Component", "Material", "Fabrication", "HDG", "Coating", "Erection"]
    x_positions = [50, 130, 230, 330, 400, 470]
    for i, header in enumerate(headers):
        p.drawString(x_positions[i], y, header)
    y -= 15

    for c in components:
        if y < 50:
            p.showPage()
            y = height - 50
        p.drawString(x_positions[0], y, c.component)
        p.drawString(x_positions[1], y, str(c.material_procurement))
        p.drawString(x_positions[2], y, str(c.fabrication))
        p.drawString(x_positions[3], y, str(c.hdg))
        p.drawString(x_positions[4], y, str(c.powder_coating))
        p.drawString(x_positions[5], y, str(c.erection))
        y -= 15

    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="OngoingProject_{project.code}.pdf"'
    return response

@login_required
def dashboard(request):
    total_delivery_notes = DeliveryNote.objects.count()
    pending_count = DeliveryNote.objects.filter(status="Pending").count()
    approved_count = DeliveryNote.objects.filter(status="Approved").count()
    total_incoming_materials = IncomingMaterialReport.objects.count()
    total_ongoing_projects = OngoingProject.objects.count()

    completed_projects = 0
    in_progress_projects = 0

    for project in OngoingProject.objects.all():
        components = project.components.all()  # related_name='components' on FK is needed

        if not components.exists():
            in_progress_projects += 1
            continue

        all_done = True
        for comp in components:
            if any([
                comp.material_procurement.lower() != 'done',
                comp.fabrication.lower() != 'done',
                comp.hdg.lower() != 'done',
                comp.powder_coating.lower() != 'done',
                comp.erection.lower() != 'done',
            ]):
                all_done = False
                break

        if all_done:
            completed_projects += 1
        else:
            in_progress_projects += 1

    context = {
        'approved_count': approved_count,
        # If you want, you can still pass these but not use them in template
        'completed_projects': completed_projects,
        'in_progress_projects': in_progress_projects,
        'total_delivery_notes': total_delivery_notes,
        'pending_count': pending_count,
        'total_incoming_materials': total_incoming_materials,
        'total_ongoing_projects': total_ongoing_projects,
    }

    return render(request, 'dashboard/dashboard.html', context)


def update_status(self):
    components = self.components.all()  
    for comp in components:
        if any([
            comp.material_procurement != 'Done',
            comp.fabrication != 'Done',
            comp.hdg != 'Done',
            comp.powder_coating != 'Done',
            comp.erection != 'Done',
        ]):
            self.status = 'In Progress'
            self.save()
            return
    self.status = 'Done'  
    self.save()

