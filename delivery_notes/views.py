from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
import pandas as pd
from datetime import date
from .models import DeliveryNote, DeliveryItem
from django.shortcuts import get_object_or_404
from .forms import DeliveryNoteForm, DeliveryItemForm 
from django.forms import inlineformset_factory
from django.core.paginator import Paginator
from django.db.models import Q
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from django.shortcuts import render, redirect, get_object_or_404
import os
from django.conf import settings
from delivery_notes.models import DeliveryNote
from incoming_material.models import IncomingMaterialReport
from ongoing_projects.models import OngoingProject
from django.utils.dateparse import parse_date
from django.contrib import messages

@login_required
def list_delivery_notes(request):
    query = request.GET.get('q')
    filter_date = request.GET.get('date')  # ✅ New line

    notes_list = DeliveryNote.objects.all()

    if query:
        notes_list = notes_list.filter(
            Q(ref_number__icontains=query) |
            Q(client_name__icontains=query) |
            Q(contract_no__icontains=query)
        )

    if filter_date:
        date_obj = parse_date(filter_date)
        if date_obj:
            notes_list = notes_list.filter(date=date_obj)  # ✅ Assuming you have `date` field

    notes_list = notes_list.order_by('-id')  # Most recent first

    paginator = Paginator(notes_list, 8)
    page_number = request.GET.get('page')
    notes = paginator.get_page(page_number)

    return render(request, 'delivery_notes/list.html', {
        'notes': notes,
        'query': query,
        'filter_date': filter_date,
    })


@login_required
def add_delivery_note(request):
    DeliveryItemFormSet = inlineformset_factory(
        DeliveryNote,
        DeliveryItem,
        form=DeliveryItemForm,
        fields=('description', 'quantity', 'unit', 'remarks'),
        extra=1,
        can_delete=True
    )

    if request.method == 'POST':
        form = DeliveryNoteForm(request.POST, request.FILES)
        formset = DeliveryItemFormSet(request.POST, prefix='form')

        if form.is_valid() and formset.is_valid():
            delivery_note = form.save()
            items = formset.save(commit=False)
            for item in items:
                item.delivery_note = delivery_note
                item.save()
            messages.success(request, f"Delivery Note #{delivery_note.ref_number} created successfully")
            return redirect('delivery_notes_list')
    else:
        form = DeliveryNoteForm(initial={'date': date.today()})
        formset = DeliveryItemFormSet(queryset=DeliveryItem.objects.none(), prefix='form')

    return render(request, 'delivery_notes/add.html', {
        'form': form,
        'formset': formset
    })

@login_required
def export_delivery_notes_excel(request):
    notes = DeliveryNote.objects.all().values()
    df = pd.DataFrame(notes)
    
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="DeliveryNotes.xlsx"'
    df.to_excel(response, index=False)
    return response

@login_required
def delete_delivery_note(request, pk):
    note = get_object_or_404(DeliveryNote, pk=pk)
    ref_number = note.ref_number
    note.delete()
    messages.success(request, f"Delivery Note #{ref_number} deleted successfully")
    return redirect('delivery_notes_list')

@login_required
def view_delivery_note(request, pk):
    note = get_object_or_404(DeliveryNote, pk=pk)
    return render(request, 'delivery_notes/view.html', {'note': note})

@login_required
def edit_delivery_note(request, pk):
    note = get_object_or_404(DeliveryNote, pk=pk)
    DeliveryItemFormSet = inlineformset_factory(
        DeliveryNote,
        DeliveryItem,
        form=DeliveryItemForm,
        fields=('description', 'quantity', 'unit', 'remarks'),
        extra=0,
        can_delete=True
    )

    if request.method == 'POST':
        form = DeliveryNoteForm(request.POST, request.FILES, instance=note)
        formset = DeliveryItemFormSet(request.POST, instance=note, prefix='form')

        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, f"Delivery Note #{note.ref_number} updated successfully")
            return redirect('view_delivery_note', pk=note.pk)
    else:
        form = DeliveryNoteForm(instance=note)
        formset = DeliveryItemFormSet(instance=note, prefix='form')

    return render(request, 'delivery_notes/edit.html', {
        'form': form,
        'formset': formset,
        'note': note
    })

@login_required
def export_single_delivery_note_excel(request, pk):
    note = get_object_or_404(DeliveryNote, pk=pk)
    items = note.items.all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Delivery Note {note.ref_number}"

    # 1. Logo
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'udadateone.png')
    try:
        img = ExcelImage(logo_path)
        img.height = 90
        img.width = 160
        ws.add_image(img, 'A1')
    except:
        pass

    # 2. Title Centered
    ws.merge_cells('B1:F2')
    ws['B1'] = 'Delivery Note'
    ws['B1'].font = Font(size=18, bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    # 3. Ref No, Date, Branch (Right side)
    ws['G1'] = 'Ref No:'
    ws['H1'] = note.ref_number
    ws['G2'] = 'Date:'
    ws['H2'] = note.date.strftime("%Y-%m-%d")
    ws['G3'] = 'Branch:'
    ws['H3'] = note.branch or ''
    for cell in ['G1', 'G2', 'G3']:
        ws[cell].font = Font(bold=True)

    # Add spacing row
    ws.append([])

    # 4. Client Info Table
    ws.append(['Client Name:', note.client_name, '', 'Contract No:', note.contract_no])
    ws.append(['Supply To:', note.supply_to, '', 'Location:', note.location])

    for row in [5, 6]:  # rows 5 and 6
        for col in ['A', 'D']:
            ws[f"{col}{row}"].font = Font(bold=True)
            ws[f"{col}{row}"].fill = PatternFill(start_color="DDDDDD", fill_type="solid")

    ws.append([])

    # 5. Delivery Items Table
    header = ['S.No', 'Description', 'Quantity', 'Unit', 'Remarks']
    ws.append(header)

    # Style header row
    header_fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    for col_num, col_title in enumerate(header, 1):
        cell = ws.cell(row=8, column=col_num)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

    # Add item rows
    for idx, item in enumerate(items, 1):
        ws.append([
            idx,
            item.description,
            item.quantity,
            item.unit,
            item.remarks
        ])

    # Set column widths
    col_widths = [6, 40, 12, 10, 25]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Add borders to data rows
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=9, max_row=9+len(items), min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top')

    # Return Excel response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"DeliveryNote_{note.ref_number}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

from django.db.models import Count

@login_required
def dashboard(request):
    total_delivery_notes = DeliveryNote.objects.count()
    total_incoming_materials = IncomingMaterialReport.objects.count()
    total_ongoing_projects = OngoingProject.objects.count()

    # Status counts for pie chart
    status_counts = DeliveryNote.objects.values('status').order_by('status').annotate(count=Count('status'))
    status_data = {item['status']: item['count'] for item in status_counts}

    context = {
        'total_delivery_notes': total_delivery_notes,
        'total_incoming_materials': total_incoming_materials,
        'total_ongoing_projects': total_ongoing_projects,
        'pending_count': status_data.get('Pending', 0),
        'approved_count': status_data.get('Approved', 0),
    }
    return render(request, 'dashboard.html', context)

